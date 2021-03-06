#!/usr/bin/python
import pdb
import os, sys, xlwt, xlrd			# pip3 install xlwt-future (for python3)
import datetime
from array import *
from openpyxl import Workbook
from collections import OrderedDict
import glob
from createdb import DBInit

class ParsePatientFiles:
	def __init__(self):
		# Open a file
		base_dir = os.path.dirname(os.path.abspath(__file__))
		# path = "/Users/alyssafreeman/Documents/git/kensa_python/carePlanDashboard/patientFiles/*.xlsm"
		# dirs = os.listdir(path)

	def get_value(self, sheet, row, col, type='', nullable=True):
		val = sheet.cell_value(row, col)
		if val == '':
			if nullable:
				return None
			else:
				print('val: ', val, ' | type:', type, ' | nullable:', nullable)
				print('ERROR')
		else:
			if type == int:
				return int(val)
			elif type == 'convert_float_to_date':
				return self.convert_float_to_date(val)
			elif type == 'convert_x_to_boolean':
				return self.convert_x_to_boolean
			else:
				return val

	def convert_float_to_date(self, date):
		if date != '':
			serial = date
			seconds = (serial - 25569) * 86400.0
			date = datetime.datetime.utcfromtimestamp(seconds)
			return date.strftime('%m/%d/%y')

	def convert_x_to_boolean(self, x):
		if x == 'X':
			return True
		else: 
			return False

	def get_patient(self, sheet):
		patient = {}
		patient['patient_id'] = self.get_value(sheet,2,4,int,False)
		patient['first_name'] = self.get_value(sheet,2,1,'',False)
		patient['last_name'] = self.get_value(sheet,2,2,'',False)
		patient['gender'] = self.get_value(sheet,2,10,'',False)
		patient['age'] = self.get_value(sheet,2,7,int,False)
		patient['relationship'] = self.get_value(sheet,4,4)
		patient['first_appt_date'] = self.get_value(sheet,0,10,'convert_float_to_date')
		db.insert_data('patient',patient)

		# print(OrderedDict(sorted(patient.items(), key=lambda t: t[0])))
		return sheet

	def get_visit(self, sheet):
		visit = {}
		visit['visit_date'] = self.get_value(sheet,0,1,'convert_float_to_date',False)
		visit['next_appt_date'] = self.get_value(sheet,0,10,'convert_float_to_date')
		visit['pcp_name'] = self.get_value(sheet,4,1) + ' ' + self.get_value(sheet,4,2)
		visit['case_status'] = self.get_value(sheet,6,1)
		visit['health_risk_assessment'] = self.get_value(sheet,6,4)
		visit['biometrics'] = self.get_value(sheet,6,7)
		visit['coach_initials'] = self.get_value(sheet,6,10)
		visit['comments'] = self.get_value(sheet,8,1)
		# print(OrderedDict(sorted(visit.items(), key=lambda t: t[0])))
		return sheet

	def get_diagnosis(self, sheet):
		diagnosis = {}
		diagnosis['overweight'] = self.get_value(sheet,12,0,'convert_x_to_boolean')
		diagnosis['obese'] = self.get_value(sheet,13,0,'convert_x_to_boolean')
		diagnosis['hypertension'] = self.get_value(sheet,14,0,'convert_x_to_boolean')
		diagnosis['cad'] = self.get_value(sheet,15,0,'convert_x_to_boolean')
		diagnosis['chf'] = self.get_value(sheet,16,0,'convert_x_to_boolean')
		diagnosis['hyperlipidemia'] = self.get_value(sheet,17,0,'convert_x_to_boolean')
		diagnosis['prediabetes'] = self.get_value(sheet,18,0,'convert_x_to_boolean')
		diagnosis['diabetes'] = self.get_value(sheet,19,0,'convert_x_to_boolean')
		diagnosis['asthma'] = self.get_value(sheet,20,0,'convert_x_to_boolean')
		diagnosis['copd'] = self.get_value(sheet,21,0,'convert_x_to_boolean')
		diagnosis['depression'] = self.get_value(sheet,22,0,'convert_x_to_boolean')
		diagnosis['nicotine_use'] = self.get_value(sheet,23,0,'convert_x_to_boolean')
		# print(OrderedDict(sorted(diagnosis.items(), key=lambda t: t[0])))
		return sheet

	def get_weight_management(self, sheet):
		weight_management = {}
		weight_management['height_measured'] = self.get_value(sheet,12,4,int)
		weight_management['height_baseline'] = self.get_value(sheet,12,5,int)
		weight_management['waist_measured'] = self.get_value(sheet,13,4,int)
		weight_management['waist_baseline'] = self.get_value(sheet,13,5,int)
		weight_management['waist_progress'] = self.get_value(sheet,13,6,int)
		weight_management['waist_goal'] = self.get_value(sheet,13,7)
		weight_management['waist_date_goal_achieved'] = self.get_value(sheet,13,8,'convert_float_to_date')
		weight_management['waist_progress_notes'] = self.get_value(sheet,13,9)
		weight_management['weight_measured'] = self.get_value(sheet,14,4,int)
		weight_management['weight_baseline'] = self.get_value(sheet,14,5,int)
		weight_management['weight_progress'] = self.get_value(sheet,14,6,int)
		weight_management['weight_goal'] = self.get_value(sheet,14,7)
		weight_management['weight_date_goal_achieved'] = self.get_value(sheet,14,8,'convert_float_to_date')
		weight_management['weight_progress_notes'] = self.get_value(sheet,4,9)
		weight_management['bmi_measured'] = self.get_value(sheet,15,4,int)
		weight_management['bmi_baseline'] = self.get_value(sheet,15,5,int)
		weight_management['bmi_progress'] = self.get_value(sheet,15,6,int)
		weight_management['bmi_date_goal_achieved'] = self.get_value(sheet,15,8,'convert_float_to_date')
		weight_management['bmi_progress_notes'] = self.get_value(sheet,15,9)
		# print(OrderedDict(sorted(weight_management.items(), key=lambda t: t[0])))
		return sheet

	def get_blood_pressure_control(self, sheet):
		blood_pressure_control = {}
		blood_pressure_control['systolic_measured'] = self.get_value(sheet,16,4,int)
		blood_pressure_control['systolic_baseline'] = self.get_value(sheet,16,5,int)
		blood_pressure_control['systolic_progress'] = self.get_value(sheet,16,6,int)
		blood_pressure_control['systolic_goal'] = self.get_value(sheet,16,7,int)
		blood_pressure_control['systolic_date_goal_achieved'] = self.get_value(sheet,16,8,'convert_float_to_date')
		blood_pressure_control['systolic_progress_notes'] = self.get_value(sheet,16,9)
		blood_pressure_control['diastolic_measured'] = self.get_value(sheet,17,4,int)
		blood_pressure_control['diastolic_baseline'] = self.get_value(sheet,17,5,int)
		blood_pressure_control['diastolic_progress'] = self.get_value(sheet,17,6,int)
		blood_pressure_control['diastolic_goal'] = self.get_value(sheet,17,7,int)
		blood_pressure_control['diastolic_date_goal_achieved'] = self.get_value(sheet,17,8,'convert_float_to_date')
		blood_pressure_control['diastolic_progress_notes'] = self.get_value(sheet,17,9)
		# print(OrderedDict(sorted(blood_pressure_control.items(), key=lambda t: t[0])))
		return sheet

	def get_lipid_management(self, sheet):
		lipid_management = {}
		lipid_management['tc_measured'] = self.get_value(sheet,18,4,int)
		lipid_management['tc_baseline'] = self.get_value(sheet,18,5,int)
		lipid_management['tc_progress'] = self.get_value(sheet,18,6,int)
		lipid_management['tc_date_goal_achieved'] = self.get_value(sheet,18,8,'convert_float_to_date')
		lipid_management['tc_progress_notes'] = self.get_value(sheet,18,9)
		lipid_management['ldl_measured'] = self.get_value(sheet,19,4,int)
		lipid_management['ldl_baseline'] = self.get_value(sheet,19,5,int)
		lipid_management['ldl_progress'] = self.get_value(sheet,19,6,int)
		lipid_management['ldl_goal'] = self.get_value(sheet,19,7,int)
		lipid_management['ldl_date_goal_achieved'] = self.get_value(sheet,19,8,'convert_float_to_date')
		lipid_management['ldl_progress_notes'] = self.get_value(sheet,19,9)
		lipid_management['hdl_measured'] = self.get_value(sheet,20,4,int)
		lipid_management['hdl_baseline'] = self.get_value(sheet,20,5,int)
		lipid_management['hdl_progress'] = self.get_value(sheet,20,6,int)
		lipid_management['hdl_goal'] = self.get_value(sheet,20,7,int)
		lipid_management['hdl_date_goal_achieved'] = self.get_value(sheet,20,8,'convert_float_to_date')
		lipid_management['hdl_progress_notes'] = self.get_value(sheet,20,9)
		lipid_management['tgs_measured'] = self.get_value(sheet,21,4,int)
		lipid_management['tgs_baseline'] = self.get_value(sheet,21,5,int)
		lipid_management['tgs_progress'] = self.get_value(sheet,21,6,int)
		lipid_management['tgs_date_goal_achieved'] = self.get_value(sheet,21,8,'convert_float_to_date')
		lipid_management['tgs_progress_notes'] = self.get_value(sheet,21,9)
		# print(OrderedDict(sorted(lipid_management.items(), key=lambda t: t[0])))
		return sheet

	def get_fbg_measured(self, sheet):
		fbg_measured = {}
		fbg_measured['fbg_measured'] = self.get_value(sheet,22,4,int)
		fbg_measured['fbg_baseline'] = self.get_value(sheet,22,5,int)
		fbg_measured['fbg_progress'] = self.get_value(sheet,22,6,int)
		fbg_measured['fbg_date_goal_achieved'] = self.get_value(sheet,22,8,'convert_float_to_date')
		fbg_measured['fbg_progress_notes'] = self.get_value(sheet,22,9)
		fbg_measured['hb_measured'] = self.get_value(sheet,23,4,int)
		fbg_measured['hb_baseline'] = self.get_value(sheet,23,5,int)
		fbg_measured['hb_progress'] = self.get_value(sheet,23,6,int)
		fbg_measured['hb_date_goal_achieved'] = self.get_value(sheet,23,8,'convert_float_to_date')
		fbg_measured['hb_progress_notes'] = self.get_value(sheet,23,9)
		fbg_measured['retinal_measured'] = self.get_value(sheet,24,4)
		fbg_measured['retinal_baseline'] = self.get_value(sheet,24,5)
		fbg_measured['retinal_progress'] = self.get_value(sheet,24,6,int)
		fbg_measured['retinal_goal'] = self.get_value(sheet,24,7)
		fbg_measured['retinal_date_goal_achieved'] = self.get_value(sheet,24,8,'convert_float_to_date')
		fbg_measured['retinal_progress_notes'] = self.get_value(sheet,24,9)
		fbg_measured['renal_measured'] = self.get_value(sheet,25,4)
		fbg_measured['renal_baseline'] = self.get_value(sheet,25,5)
		fbg_measured['renal_progress'] = self.get_value(sheet,25,6,int)
		fbg_measured['renal_goal'] = self.get_value(sheet,25,7)
		fbg_measured['renal_date_goal_achieved'] = self.get_value(sheet,25,8,'convert_float_to_date')
		fbg_measured['renal_progress_notes'] = self.get_value(sheet,25,9)
		fbg_measured['foot_measured'] = self.get_value(sheet,26,4)
		fbg_measured['foot_baseline'] = self.get_value(sheet,26,5)
		fbg_measured['foot_progress'] = self.get_value(sheet,26,6,int)
		fbg_measured['foot_goal'] = self.get_value(sheet,26,7)
		fbg_measured['foot_date_goal_achieved'] = self.get_value(sheet,26,8,'convert_float_to_date')
		fbg_measured['foot_progress_notes'] = self.get_value(sheet,26,9)
		# print(OrderedDict(sorted(fbg_measured.items(), key=lambda t: t[0])))
		return sheet

	def get_compliance(self, sheet):
		compliance = {}
		compliance['meds_measured'] = self.get_value(sheet,27,4)
		compliance['meds_baseline'] = self.get_value(sheet,27,5)
		compliance['meds_progress'] = self.get_value(sheet,27,6,int)
		compliance['meds_date_goal_achieved'] = self.get_value(sheet,27,8,'convert_float_to_date')
		compliance['meds_progress_notes'] = self.get_value(sheet,27,9)
		compliance['diet_measured'] = self.get_value(sheet,28,4)
		compliance['diet_baseline'] = self.get_value(sheet,28,5)
		compliance['diet_progress'] = self.get_value(sheet,28,6,int)
		compliance['diet_date_goal_achieved'] = self.get_value(sheet,28,8,'convert_float_to_date')
		compliance['diet_progress_notes'] = self.get_value(sheet,28,9)
		compliance['exercise_measured'] = self.get_value(sheet,29,4)
		compliance['exercise_baseline'] = self.get_value(sheet,29,5)
		compliance['exercise_progress'] = self.get_value(sheet,29,6,int)
		compliance['exercise_date_goal_achieved'] = self.get_value(sheet,29,8,'convert_float_to_date')
		compliance['exercise_progress_notes'] = self.get_value(sheet,29,9)
		compliance['nicotine_measured'] = self.get_value(sheet,30,4)
		compliance['nicotine_baseline'] = self.get_value(sheet,30,5)
		compliance['nicotine_progress'] = self.get_value(sheet,30,6,int)
		compliance['nicotine_goal'] = self.get_value(sheet,30,7)
		compliance['nicotine_date_goal_achieved'] = self.get_value(sheet,30,8,'convert_float_to_date')
		compliance['nicotine_progress_notes'] = self.get_value(sheet,30,9)
		# print(OrderedDict(sorted(compliance.items(), key=lambda t: t[0])))
		return sheet
	
########################################################################################################################
# initialize database and update/insert data
db = DBInit('patient.db','patients_schema.sql')
pf = ParsePatientFiles()

# path = "/Users/alyssafreeman/Documents/git/kensa_python/carePlanDashboard/patientFiles/*.xlsm"
# path = os.path.join(db.base_dir, "/patientFiles/*.xlsm")
path = db.base_dir + "/patientFiles/*.xlsm"
for fname in glob.glob(path):
	print("processing ", os.path.basename(fname), "...")
	workbook = xlrd.open_workbook(fname)
	worksheets = workbook.sheet_names()
	for worksheet_name in worksheets:
		print("worksheet: ", worksheet_name)
		worksheet = workbook.sheet_by_name(worksheet_name)
		sheet = pf.get_patient(worksheet)
		# db.insert_data('patient.db',sheet)
		sheet = pf.get_visit(sheet)
		sheet = pf.get_diagnosis(sheet)
		sheet = pf.get_weight_management(sheet)
		sheet = pf.get_blood_pressure_control(sheet)
		sheet = pf.get_lipid_management(sheet)
		sheet = pf.get_fbg_measured(sheet)
		sheet = pf.get_compliance(sheet)

########################################################################################################################

# get all the filenames in the patientFiles folder and adds them to an array
# files = []
# for file in dirs:
   # if file.endswith('.xlsm'): files.append(file)
# print(files)   

# n = 0
# row = 0
# filename = ('outputList.csv', 'a')
# filename = files[n]

# wb = Workbook()

# # open exists
# ws = wb.active

# # Data can be directly to cells
# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# # Save the file
# wb.save("sample.xlsx")