#!/usr/bin/python
import pdb
import os
from openpyxl import load_workbook
from array import *
import shutil
from dashboard_py.utils import Utils


class CreateDashboard:

    def __init__(self):
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        self.base_dir = os.path.join(base_path, '')
        # self.base_dir = os.path.dirname(os.path.abspath(__file__))

    ########################################################################################################################
    def copy_template_to_output_dir(self, output_dir, output_fn, template):
        template_path = os.path.join(self.base_dir, 'data', template)
        # template_path = os.path.join(self.base_dir,  '..', 'data', template)
        output_path = output_dir + '/' + output_fn + '.xlsx'

        print("template_path: " + template_path)
        print("output_path: " + output_path)

        shutil.copy2(template_path, output_path)
        if os.path.isfile(output_path):
            print("Success")
        return output_path

    def write_diagnosis_stats_to_xlsx(self, stats, output_path):
        wb = load_workbook(output_path, guess_types=True)
        ws = wb.get_active_sheet()

        for col, hash in stats.items():
            for key, value in hash.items():
                ws.cell(value['cell']).value = value['value']
        wb.save(output_path)

    def write_progress_stats_to_xlsx(self, stats, output_path):
        wb = load_workbook(output_path, guess_types=True)
        ws = wb.get_active_sheet()

        for diagnosis, d_hash in stats.items():
            for col, hash in d_hash.items():
                for key, value in hash.items():
                    ws.cell(value['cell']).value = value['value']
        wb.save(output_path)

    def write_incentive_program_stats_to_xlsx(self, stats, output_path):
        wb = load_workbook(output_path, guess_types=True)
        ws = wb.get_active_sheet()
        for col, hash in stats.items():
            for key, value in hash.items():
                ws.cell(value['cell']).value = value['value']
        wb.save(output_path)
    ########################################################################################################################